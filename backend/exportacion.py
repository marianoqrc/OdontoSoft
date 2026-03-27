import os
import hashlib
from datetime import datetime
from openpyxl import Workbook, load_workbook
from config import DATA_DIR

SALT = "odontosoft2024"

def anonimizar(valor):
    """Reemplaza un valor sensible por un hash SHA-256 corto."""
    if not valor:
        return ""
    return hashlib.sha256(f"{SALT}{valor}".encode()).hexdigest()[:12]

def consolidar(anonimizar_datos=False):
    """
    RF07 — Une todos los pacientes en un dataset maestro.
    RF02 — Si anonimizar=True, reemplaza nombre/dni por hashes.
    Devuelve (wb, cantidad_pacientes, cantidad_eventos)
    """
    if not os.path.exists(DATA_DIR):
        return None, 0, 0

    wb = Workbook()

    # ── Hoja 1: Pacientes ─────────────────────────────────────────────────
    ws_pac = wb.active
    ws_pac.title = "Pacientes"
    ws_pac.append([
        "id_paciente", "nombre", "apellido", "dni",
        "fecha_nacimiento", "telefono", "email",
        "obra_social", "nro_afiliado", "alergias", "creado_en"
    ])

    # Estilo encabezado
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    for cell in ws_pac[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Hoja 2: Historia clínica ──────────────────────────────────────────
    ws_hist = wb.create_sheet("Historia")
    ws_hist.append([
        "id_paciente", "fecha", "piezas",
        "procedimiento", "descripcion"
    ])

    for cell in ws_hist[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_pacientes = 0
    total_eventos = 0

    for dni in os.listdir(DATA_DIR):
        ruta = os.path.join(DATA_DIR, str(dni), "paciente.xlsx")
        if not os.path.exists(ruta):
            continue

        try:
            source_wb = load_workbook(ruta, read_only=True)

            # Leer datos del paciente
            ws_p = source_wb["Paciente"]
            headers_p = [c.value for c in ws_p[1]]
            if ws_p.max_row < 2:
                source_wb.close()
                continue

            # Buscar última fila con datos
            ultima = None
            for row in ws_p.iter_rows(min_row=2):
                vals = [c.value for c in row]
                if any(v is not None for v in vals):
                    ultima = vals
            if not ultima:
                source_wb.close()
                continue

            paciente = dict(zip(headers_p, ultima))
            if paciente.get("activo") == False:
                source_wb.close()
                continue

            # ID anonimizado o DNI real
            id_pac = anonimizar(str(paciente.get("dni", ""))) if anonimizar_datos else paciente.get("dni", "")
            nombre  = "ANONIMIZADO" if anonimizar_datos else paciente.get("nombre", "")
            apellido = "ANONIMIZADO" if anonimizar_datos else paciente.get("apellido", "")
            dni_export = anonimizar(str(paciente.get("dni", ""))) if anonimizar_datos else paciente.get("dni", "")

            ws_pac.append([
                id_pac, nombre, apellido, dni_export,
                paciente.get("fecha_nacimiento", ""),
                "" if anonimizar_datos else paciente.get("telefono", ""),
                "" if anonimizar_datos else paciente.get("email", ""),
                paciente.get("obra_social", ""),
                paciente.get("nro_afiliado", ""),
                paciente.get("alergias", ""),
                paciente.get("creado_en", ""),
            ])
            total_pacientes += 1

            # Leer historia clínica
            if "Historia" in source_wb.sheetnames:
                ws_h = source_wb["Historia"]
                headers_h = [c.value for c in ws_h[1]]

                for row in ws_h.iter_rows(min_row=2):
                    vals = [c.value for c in row]
                    if not any(v is not None for v in vals):
                        continue
                    evento = dict(zip(headers_h, vals))
                    ws_hist.append([
                        id_pac,
                        evento.get("fecha", ""),
                        evento.get("piezas", ""),
                        evento.get("procedimiento", ""),
                        evento.get("descripcion", ""),
                    ])
                    total_eventos += 1

            source_wb.close()

        except Exception as e:
            print(f"Error procesando paciente {dni}: {e}")

    # Ajustar ancho de columnas automáticamente
    for ws in [ws_pac, ws_hist]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    return wb, total_pacientes, total_eventos