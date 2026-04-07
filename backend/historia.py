import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from config import DATA_DIR
from werkzeug.utils import secure_filename # Asegurate de importar esto o instalar werkzeug

def get_ruta_paciente(dni):
    return os.path.join(DATA_DIR, str(dni), "paciente.xlsx")

def get_carpeta_paciente(dni):
    return os.path.join(DATA_DIR, str(dni))

def leer_historia(dni):
    """Lee todos los eventos de la historia clínica."""
    ruta = get_ruta_paciente(dni)
    if not os.path.exists(ruta):
        return []

    try:
        wb = load_workbook(ruta, read_only=True)
        ws = wb["Historia"]
        headers = [cell.value for cell in ws[1]]
        eventos = []

        for row in ws.iter_rows(min_row=2):
            valores = [cell.value for cell in row]
            if any(v is not None for v in valores):
                eventos.append(dict(zip(headers, valores)))

        wb.close()
        # Más reciente primero
        return list(reversed(eventos))

    except Exception as e:
        print(f"Error leyendo historia {dni}: {e}")
        return []

def agregar_evento(dni, evento):
    """Agrega un evento a la historia clínica — nunca edita los anteriores."""
    ruta = get_ruta_paciente(dni)
    if not os.path.exists(ruta):
        raise ValueError("Paciente no encontrado")

    try:
        wb = load_workbook(ruta)
        ws = wb["Historia"]

        # Timestamp automático (RF06)
        id_adjunto = datetime.now().strftime("%Y-%m-%d_%H%M%S")

        nuevo = {
            "fecha":      datetime.now().isoformat(),
            "piezas":     ",".join(str(p) for p in evento.get("piezas", [])),
            "procedimiento": evento.get("procedimiento", ""),
            "descripcion":   evento.get("descripcion", ""),
            "profesional":   evento.get("profesional", ""),
            "id_adjunto":    id_adjunto,
        }

        headers = [cell.value for cell in ws[1]]
        fila = [nuevo.get(h, "") for h in headers]
        ws.append(fila)

        wb.save(ruta)
        
        # Devolvemos el id_adjunto generado para que la ruta de Flask lo use al guardar los archivos
        return {"ok": True, "id_adjunto": id_adjunto}

    except PermissionError:
        raise PermissionError(
            "El archivo está abierto en Excel. Cerralo y volvé a intentar."
        )

def guardar_adjuntos(dni, id_adjunto, archivos_lista):
    """
    Recibe una lista de archivos de Flask (request.files.getlist) 
    y los guarda en una subcarpeta dentro del paciente.
    """
    if not archivos_lista:
        return
    
    # Crea la ruta: DATA_DIR / dni / id_adjunto /
    carpeta_destino = os.path.join(get_carpeta_paciente(dni), str(id_adjunto))
    os.makedirs(carpeta_destino, exist_ok=True)
    
    for archivo in archivos_lista:
        if archivo and archivo.filename != '':
            # Limpia el nombre del archivo por seguridad
            nombre_seguro = secure_filename(archivo.filename)
            ruta_final = os.path.join(carpeta_destino, nombre_seguro)
            archivo.save(ruta_final)