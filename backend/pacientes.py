import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from config import DATA_DIR

def get_ruta_paciente(dni):
    carpeta = os.path.join(DATA_DIR, str(dni))
    os.makedirs(carpeta, exist_ok=True)
    return os.path.join(carpeta, "paciente.xlsx")

def crear_workbook_nuevo():
    wb = Workbook()
    ws = wb.active
    ws.title = "Paciente"
    ws.append([
        "dni", "nombre", "apellido", "fecha_nacimiento",
        "telefono", "email", "obra_social", "nro_afiliado",
        "hipertenso", "diabetico", "alteracion_coagulacion",
        "fuma", "tiempo_fumador", "asma", "consume_drogas", "detalles_drogas",
        "insuficiencia_renal", "insuficiencia_hepatica",
        "embarazada", "tiempo_embarazo", "valor_presion",
        "alergias", "activo", "creado_en"
    ])
    ws_historia = wb.create_sheet("Historia")
    ws_historia.append([
        "fecha", "piezas", "procedimiento", "descripcion", "profesional", "id_adjunto"
    ])
    return wb

def listar_pacientes():
    if not os.path.exists(DATA_DIR):
        return []

    pacientes = []

    for dni in os.listdir(DATA_DIR):
        # Ignorar archivos, solo procesar carpetas
        ruta_carpeta = os.path.join(DATA_DIR, dni)
        if not os.path.isdir(ruta_carpeta):
            continue

        ruta = get_ruta_paciente(dni)
        if not os.path.exists(ruta):
            continue
        try:
            wb = load_workbook(ruta, read_only=True)
            ws = wb["Paciente"]
            headers = [cell.value for cell in ws[1]]
            if ws.max_row < 2:
                wb.close()
                continue

            ultima_fila = None
            for row in ws.iter_rows(min_row=2):
                valores = [cell.value for cell in row]
                if any(v is not None for v in valores):
                    ultima_fila = valores

            wb.close()

            if not ultima_fila:
                continue

            paciente = dict(zip(headers, ultima_fila))

            if paciente.get("activo") != False and paciente.get("dni"):
                pacientes.append(paciente)

        except Exception as e:
            print(f"Error leyendo paciente {dni}: {e}")

    return sorted(pacientes, key=lambda p: p.get("apellido") or "")

def guardar_paciente(datos):
    print("Datos recibidos:", datos)
    dni = datos.get("dni")
    if not dni:
        raise ValueError("El DNI es obligatorio")
        
    # NUEVA VALIDACIÓN: El teléfono ahora es obligatorio
    telefono = datos.get("telefono")
    if not telefono or str(telefono).strip() == "":
        raise ValueError("El teléfono es obligatorio para enviar recordatorios por WhatsApp")

    ruta = get_ruta_paciente(dni)

    try:
        if os.path.exists(ruta):
            wb = load_workbook(ruta)
            ws = wb["Paciente"]
            # Borrar TODAS las filas de datos (fila 2 en adelante)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            # Escribir en fila 2 directamente
            headers = [cell.value for cell in ws[1]]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=2, column=col_idx, value=datos.get(header))
        else:
            wb = crear_workbook_nuevo()
            ws = wb["Paciente"]
            datos["creado_en"] = datetime.now().isoformat()
            datos["activo"] = True
            headers = [cell.value for cell in ws[1]]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=2, column=col_idx, value=datos.get(header))

        wb.save(ruta)
        return {"ok": True, "dni": dni}

    except PermissionError:
        raise PermissionError(
            f"El archivo del paciente está abierto en Excel. "
            f"Cerralo y volvé a intentar."
        )
    
def dar_de_baja(dni):
    if not dni or dni == "null":
        return {"ok": False, "error": "DNI inválido"}

    ruta = get_ruta_paciente(str(dni))
    if not os.path.exists(ruta):
        return {"ok": False, "error": "Paciente no encontrado"}

    try:
        wb = load_workbook(ruta)
        ws = wb["Paciente"]
        headers = [cell.value for cell in ws[1]]

        if ws.max_row < 2:
            return {"ok": False, "error": "Paciente sin datos"}

        # Leer datos actuales de la última fila con datos
        ultima_fila = None
        for row in ws.iter_rows(min_row=2):
            valores = [cell.value for cell in row]
            if any(v is not None for v in valores):
                ultima_fila = valores

        if not ultima_fila:
            return {"ok": False, "error": "Paciente sin datos"}

        paciente = dict(zip(headers, ultima_fila))
        paciente["activo"] = False

        # Borrar y reescribir
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col_idx, value=paciente.get(header))

        wb.save(ruta)
        return {"ok": True}

    except PermissionError:
        raise PermissionError("El archivo está abierto en Excel. Cerralo y volvé a intentar.")