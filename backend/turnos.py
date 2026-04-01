import os
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from config import DATA_DIR
import pacientes as pac

TURNOS_FILE = os.path.join(DATA_DIR, '_turnos.xlsx')

HORA_MANANA_INICIO = "09:00"
HORA_MANANA_FIN    = "13:00"
HORA_TARDE_INICIO  = "14:00"
HORA_TARDE_FIN     = "20:00"
DURACION_MODULO    = 15  # minutos

def generar_modulos(hora_inicio, hora_fin):
    """Genera lista de módulos de 15 min entre dos horas."""
    modulos = []
    h, m = map(int, hora_inicio.split(':'))
    hf, mf = map(int, hora_fin.split(':'))
    fin_mins = hf * 60 + mf
    actual = h * 60 + m
    while actual < fin_mins:
        modulos.append(f"{actual//60:02d}:{actual%60:02d}")
        actual += DURACION_MODULO
    return modulos

MODULOS_MANANA = generar_modulos(HORA_MANANA_INICIO, HORA_MANANA_FIN)
MODULOS_TARDE  = generar_modulos(HORA_TARDE_INICIO,  HORA_TARDE_FIN)
TODOS_MODULOS  = MODULOS_MANANA + MODULOS_TARDE

def init_turnos():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(TURNOS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Turnos"
        ws.append([
            "id", "fecha", "hora_inicio", "hora_fin",
            "dni_paciente", "nombre_paciente", "motivo",
            "estado", "creado_en"
        ])
        wb.save(TURNOS_FILE)

def leer_todos():
    init_turnos()
    try:
        wb = load_workbook(TURNOS_FILE, read_only=True)
        ws = wb["Turnos"]
        headers = [c.value for c in ws[1]]
        turnos = []
        for row in ws.iter_rows(min_row=2):
            vals = [c.value for c in row]
            if any(v is not None for v in vals):
                turnos.append(dict(zip(headers, vals)))
        wb.close()
        return turnos
    except Exception as e:
        print(f"Error leyendo turnos: {e}")
        return []

def hora_a_mins(hora):
    h, m = map(int, hora.split(':'))
    return h * 60 + m

def mins_a_hora(mins):
    return f"{mins//60:02d}:{mins%60:02d}"

def leer_por_fecha(fecha_str):
    todos = leer_todos()
    return [t for t in todos
            if str(t.get("fecha","")) == fecha_str
            and t.get("estado") != "cancelado"]

def get_disponibilidad(fecha_str, turno="manana"):
    """
    Devuelve módulos con estado libre/ocupado para el turno seleccionado.
    """
    modulos = MODULOS_MANANA if turno == "manana" else MODULOS_TARDE
    ocupados = leer_por_fecha(fecha_str)

    # Construir set de módulos ocupados
    modulos_ocupados = {}
    for t in ocupados:
        inicio = hora_a_mins(str(t["hora_inicio"]))
        fin    = hora_a_mins(str(t["hora_fin"]))
        cur = inicio
        while cur < fin:
            modulos_ocupados[mins_a_hora(cur)] = t
            cur += DURACION_MODULO

    resultado = []
    for m in modulos:
        if m in modulos_ocupados:
            resultado.append({
                "hora":   m,
                "estado": "ocupado",
                "turno":  modulos_ocupados[m],
            })
        else:
            resultado.append({
                "hora":   m,
                "estado": "libre",
                "turno":  None,
            })
    return resultado

def get_timeline(fecha_str, turno="manana"):
    """
    Devuelve el timeline del día consolidando bloques ocupados y libres,
    incluyendo el teléfono del paciente.
    """
    modulos = MODULOS_MANANA if turno == "manana" else MODULOS_TARDE
    hora_fin_str = HORA_MANANA_FIN if turno == "manana" else HORA_TARDE_FIN
    ocupados = leer_por_fecha(fecha_str)
    
    # NUEVO: Obtenemos todos los pacientes para cruzar datos
    lista_pacientes = pac.listar_pacientes()
    mapa_pacientes = {str(p.get("dni")): p.get("telefono", "") for p in lista_pacientes}

    # Turnos del período
    inicio_periodo = hora_a_mins(modulos[0])
    fin_periodo    = hora_a_mins(hora_fin_str)

    turnos_periodo = []
    for t in ocupados:
        inicio = hora_a_mins(str(t["hora_inicio"]))
        fin    = hora_a_mins(str(t["hora_fin"]))
        
        # NUEVO: Agregamos el teléfono al turno cruzando por DNI
        dni_paciente = str(t.get("dni_paciente", ""))
        telefono = mapa_pacientes.get(dni_paciente, "")
        t["telefono_paciente"] = telefono
        
        if inicio >= inicio_periodo and fin <= fin_periodo:
            turnos_periodo.append({
                **t,
                "inicio_mins": inicio,
                "fin_mins":    fin,
            })

    turnos_periodo.sort(key=lambda x: x["inicio_mins"])

    bloques = []
    cursor = inicio_periodo

    for t in turnos_periodo:
        # Bloque libre antes del turno
        if cursor < t["inicio_mins"]:
            bloques.append({
                "tipo":        "libre",
                "hora_inicio": mins_a_hora(cursor),
                "hora_fin":    mins_a_hora(t["inicio_mins"]),
                "duracion_min": t["inicio_mins"] - cursor,
            })
        # Bloque ocupado
        bloques.append({
            "tipo":         "ocupado",
            "hora_inicio":  mins_a_hora(t["inicio_mins"]),
            "hora_fin":     mins_a_hora(t["fin_mins"]),
            "duracion_min": t["fin_mins"] - t["inicio_mins"],
            "turno":        t,
        })
        cursor = t["fin_mins"]

    # Bloque libre al final
    if cursor < fin_periodo:
        bloques.append({
            "tipo":         "libre",
            "hora_inicio":  mins_a_hora(cursor),
            "hora_fin":     mins_a_hora(fin_periodo),
            "duracion_min": fin_periodo - cursor,
        })

    return bloques

def crear_turno(datos):
    init_turnos()
    fecha       = datos.get("fecha")
    hora_inicio = datos.get("hora_inicio")
    hora_fin    = datos.get("hora_fin")

    # Verificar que los módulos estén libres
    ocupados = leer_por_fecha(fecha)
    inicio_nuevo = hora_a_mins(hora_inicio)
    fin_nuevo    = hora_a_mins(hora_fin)

    for t in ocupados:
        ini = hora_a_mins(str(t["hora_inicio"]))
        fin = hora_a_mins(str(t["hora_fin"]))
        if not (fin_nuevo <= ini or inicio_nuevo >= fin):
            raise ValueError(
                f"El horario se superpone con un turno existente "
                f"({t['hora_inicio']} - {t['hora_fin']})"
            )

    todos = leer_todos()
    nuevo_id = max((t.get("id", 0) or 0 for t in todos), default=0) + 1

    try:
        wb = load_workbook(TURNOS_FILE)
        ws = wb["Turnos"]
        ws.append([
            nuevo_id, fecha, hora_inicio, hora_fin,
            datos.get("dni_paciente", ""),
            datos.get("nombre_paciente", ""),
            datos.get("motivo", ""),
            "confirmado",
            datetime.now().isoformat(),
        ])
        wb.save(TURNOS_FILE)
        return {"ok": True, "id": nuevo_id}
    except PermissionError:
        raise PermissionError("El archivo de turnos está abierto. Cerralo y volvé a intentar.")

def cancelar_turno(turno_id):
    init_turnos()
    try:
        wb = load_workbook(TURNOS_FILE)
        ws = wb["Turnos"]
        headers = [c.value for c in ws[1]]
        id_col     = headers.index("id") + 1
        estado_col = headers.index("estado") + 1
        for row in ws.iter_rows(min_row=2):
            if row[id_col-1].value == turno_id:
                row[estado_col-1].value = "cancelado"
                break
        wb.save(TURNOS_FILE)
        return {"ok": True}
    except PermissionError:
        raise PermissionError("El archivo de turnos está abierto. Cerralo y volvé a intentar.")
    
def proximos_turnos(dias=7):
    hoy = date.today()
    todos = leer_todos()
    resultado = []
    for i in range(dias):
        dia = hoy + timedelta(days=i)
        if dia.weekday() in [0,1,2,3,4,5]:
            fecha_str = dia.isoformat()
            del_dia = [t for t in todos
                      if str(t.get("fecha","")) == fecha_str
                      and t.get("estado") != "cancelado"]
            resultado.extend(del_dia)
    return sorted(resultado, key=lambda t: (str(t["fecha"]), str(t["hora_inicio"])))

def actualizar_turno(turno_id, datos):
    init_turnos()
    try:
        wb = load_workbook(TURNOS_FILE)
        ws = wb["Turnos"]
        headers = [c.value for c in ws[1]]
        id_col = headers.index("id") + 1
        nombre_col = headers.index("nombre_paciente") + 1
        motivo_col = headers.index("motivo") + 1

        for row in ws.iter_rows(min_row=2):
            if row[id_col-1].value == turno_id:
                # Si mandamos nombre nuevo, lo actualiza
                if "nombre_paciente" in datos:
                    row[nombre_col-1].value = datos["nombre_paciente"]
                # Si mandamos motivo nuevo, lo actualiza
                if "motivo" in datos:
                    row[motivo_col-1].value = datos["motivo"]
                break
                
        wb.save(TURNOS_FILE)
        return {"ok": True}
    except PermissionError:
        raise PermissionError("El archivo de turnos está abierto. Cerralo y volvé a intentar.")